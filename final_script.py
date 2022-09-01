#libraries
import os
from tkinter import ANCHOR
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
from natsort import natsorted
import openpyxl
import barcode
from barcode.writer import ImageWriter

#function
def generate_barcode(number):
    barcode_format = barcode.get_barcode_class('code128')
    my_barcode = barcode_format(number, writer=ImageWriter())
    my_barcode.save(f"{number}")
#taking user input
a= input("Enter File name with full path: ")
# Define variable to load the wookbook
wookbook = openpyxl.load_workbook(a)

# Define variable to read the active sheet:
worksheet = wookbook.active

# Iterate the loop to read the cell values and generate image
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        barcode_final = col[i].value
        generate_barcode(barcode_final)
wookbook.close()
workbook= Workbook()
worksheet= workbook.active
#resize cell
for row in range(1,100):
    for col in range (1,3):
        worksheet.row_dimensions[row].height=198
        col_letter= get_column_letter(col)
        worksheet.column_dimensions[col_letter].width=50

#images list
images= []
for filename in natsorted(glob.glob('*.png')):
    images.append(filename)
#insert images
for index, image in enumerate(images):
    worksheet.add_image(Image(image), anchor='A'+ str(index+1))
#taking user input    
save= input("Save as: ")
workbook.save(save+".xlsx")
cmd='rm -rf *.png'
os.system(cmd)
    
